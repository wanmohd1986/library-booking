from flask import Flask, render_template, request, redirect
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

FILE = "bookings.xlsx"

if not os.path.exists(FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Nama", "Tarikh", "Masa", "Tujuan"])
    wb.save(FILE)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/book", methods=["POST"])
def book():
    nama = request.form["nama"]
    tarikh = request.form["tarikh"]
    masa = request.form["masa"]
    tujuan = request.form["tujuan"]

    wb = load_workbook(FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == tarikh and row[2] == masa:
            return "Slot sudah ditempah!"

    ws.append([nama, tarikh, masa, tujuan])
    wb.save(FILE)

    return redirect("/list")

@app.route("/list")
def list_booking():
    wb = load_workbook(FILE)
    ws = wb.active
    data = list(ws.iter_rows(min_row=2, values_only=True))
    return render_template("list.html", data=data)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)